"""
Draft Benchmark Runner — runs 10 civil scenarios, fills Draft_test.xlsx.

Usage:
    agent_steer/Scripts/python.exe research/run_draft_benchmark.py

Modes:
    --draft     Run pipeline on all 10 scenarios, fill "Draft Agent" column (default)
    --compare   After user fills ChatGPT-5.4 column, run error analysis + comparison
"""
from __future__ import annotations

import argparse
import asyncio
import json
import re
import sys
import time
from datetime import datetime
from pathlib import Path

# Fix Windows console encoding
if sys.stdout.encoding != "utf-8":
    sys.stdout.reconfigure(encoding="utf-8")

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

EXCEL_PATH = Path(__file__).resolve().parents[1] / "docs" / "Draft_test.xlsx"
OUTPUT_DIR = Path(__file__).parent / "output"
OUTPUT_DIR.mkdir(exist_ok=True)


# ── Helpers ──────────────────────────────────────────────────────────────────

def _as_dict(val):
    if val is None:
        return {}
    if isinstance(val, dict):
        return val
    fn = getattr(val, "model_dump", None)
    return fn() if callable(fn) else {}


def _extract_draft_text(result: dict) -> str:
    """Extract draft text from pipeline result."""
    final_block = _as_dict(result.get("final_draft"))
    artifacts = final_block.get("draft_artifacts") or []
    if artifacts:
        first = artifacts[0] if isinstance(artifacts[0], dict) else _as_dict(artifacts[0])
        return (first.get("text") or "").strip()

    draft_block = _as_dict(result.get("draft"))
    draft_arts = draft_block.get("draft_artifacts") or []
    if draft_arts:
        first = draft_arts[0] if isinstance(draft_arts[0], dict) else _as_dict(draft_arts[0])
        return (first.get("text") or "").strip()
    return ""


# ── Error Analysis ───────────────────────────────────────────────────────────

ERROR_CATEGORIES = {
    "fabrication": {
        "label": "Fabricated facts/documents/citations",
        "patterns": [
            r"air\s+\d{4}\s+\w+\s+\d+",          # Fabricated AIR citations
            r"\d{4}\s+scc\s+\d+",                   # Fabricated SCC citations
            r"ilr\s+\d{4}",                          # Fabricated ILR citations
            r"annexure[-\s]+[a-z]\s*[-–—:]\s*\w+.*(?:dated|copy)", # Invented annexures
        ],
    },
    "wrong_statute": {
        "label": "Wrong/repealed statute cited",
        "patterns": [
            r"indian\s+evidence\s+act,?\s*1872",    # Repealed — now BSA 2023
            r"code\s+of\s+criminal\s+procedure,?\s*1973",  # Repealed — now BNSS 2023
            r"indian\s+penal\s+code",                # Repealed — now BNS 2023
            r"section\s+27a\s+of\s+the\s+specific\s+relief", # S.27A doesn't exist
        ],
    },
    "missing_section": {
        "label": "Missing mandatory section",
        "checks": [
            ("verification", r"verif(?:ication|ied)"),
            ("prayer", r"prayer"),
            ("court_heading", r"(?:district|civil)\s+(?:court|judge)"),
            ("parties", r"plaintiff|petitioner"),
            ("jurisdiction", r"jurisdiction|territorial|pecuniary"),
            ("cause_of_action", r"cause\s+of\s+action"),
            ("valuation", r"valuat"),
            ("court_fee", r"court\s+fee"),
        ],
    },
    "legal_error": {
        "label": "Legal reasoning error",
        "checks": [
            ("limitation_wrong_trigger", r"limitation\s+(?:period\s+)?runs\s+from[^.]{0,80}notice"),
            ("interest_wrong_cite", r"pendente\s+lite.*section\s+34\s+cpc"),
            ("and_or_usage", r"\band/or\b"),
            ("facts_law_mixing", None),  # Custom check
        ],
    },
    "placeholder_excess": {
        "label": "Excessive/unnecessary placeholders",
        "pattern": r"\{\{[A-Z_]+\}\}",
        "threshold": 15,
    },
    "structural": {
        "label": "Structural/formatting issues",
        "checks": [
            ("no_paragraph_numbers", None),  # Custom
            ("no_continuous_numbering", None),  # Custom
        ],
    },
}


def _check_facts_law_separation(text: str) -> bool:
    """Check that FACTS section doesn't contain statutory citations."""
    m = re.search(
        r"facts\s+of\s+the\s+case\s+(.*?)(?:legal\s+basis|cause\s+of\s+action)",
        text, re.DOTALL | re.IGNORECASE,
    )
    if not m:
        return True
    facts_text = m.group(1)
    return not bool(re.search(
        r"section\s+\d+\s+of\s+the\s+(?:indian|contract|evidence|limitation|specific|civil|transfer|partnership)",
        facts_text, re.IGNORECASE,
    ))


def analyze_errors(text: str, scenario_num: int) -> dict:
    """Analyze a draft for errors across all categories."""
    t_lower = text.lower()
    errors = []

    # 1. Fabrication check
    for pat in ERROR_CATEGORIES["fabrication"]["patterns"]:
        matches = re.findall(pat, t_lower)
        if matches:
            errors.append({
                "category": "fabrication",
                "severity": "critical",
                "detail": f"Fabricated citation/document: {matches[:3]}",
            })

    # 2. Wrong statute check
    for pat in ERROR_CATEGORIES["wrong_statute"]["patterns"]:
        matches = re.findall(pat, t_lower)
        if matches:
            errors.append({
                "category": "wrong_statute",
                "severity": "critical",
                "detail": f"Wrong/repealed statute: {matches[0]}",
            })

    # 3. Missing sections
    for section_name, pattern in ERROR_CATEGORIES["missing_section"]["checks"]:
        if not re.search(pattern, t_lower):
            errors.append({
                "category": "missing_section",
                "severity": "high",
                "detail": f"Missing: {section_name}",
            })

    # 4. Legal errors
    if re.search(r"limitation\s+(?:period\s+)?runs\s+from[^.]{0,80}notice", t_lower):
        errors.append({
            "category": "legal_error",
            "severity": "high",
            "detail": "Limitation anchored to notice date (usually wrong)",
        })
    if re.search(r"pendente\s+lite.*section\s+34\s+cpc", t_lower):
        errors.append({
            "category": "legal_error",
            "severity": "medium",
            "detail": "Pendente lite interest cites S.34 CPC (should be Order XX Rule 11 CPC)",
        })
    if re.search(r"\band/or\b", t_lower):
        errors.append({
            "category": "legal_error",
            "severity": "low",
            "detail": "'and/or' used — courts disfavor this",
        })
    if not _check_facts_law_separation(text):
        errors.append({
            "category": "legal_error",
            "severity": "medium",
            "detail": "Facts section contains statutory citations (should be in Legal Basis)",
        })

    # 5. Placeholder excess
    placeholders = re.findall(r"\{\{[A-Z_]+\}\}", text)
    if len(placeholders) > ERROR_CATEGORIES["placeholder_excess"]["threshold"]:
        errors.append({
            "category": "placeholder_excess",
            "severity": "medium",
            "detail": f"{len(placeholders)} placeholders (threshold: 15)",
        })

    # 6. Structural checks
    para_nums = re.findall(r"(?:^|\n)\s*(\d+)\.\s", text)
    if len(para_nums) < 5:
        errors.append({
            "category": "structural",
            "severity": "medium",
            "detail": f"Only {len(para_nums)} numbered paragraphs (expected 10+)",
        })
    else:
        # Check continuous numbering
        nums = [int(n) for n in para_nums]
        expected = list(range(1, len(nums) + 1))
        if nums != expected:
            errors.append({
                "category": "structural",
                "severity": "low",
                "detail": f"Paragraph numbering not continuous: {nums[:10]}...",
            })

    # Score
    critical = sum(1 for e in errors if e["severity"] == "critical")
    high = sum(1 for e in errors if e["severity"] == "high")
    medium = sum(1 for e in errors if e["severity"] == "medium")
    low = sum(1 for e in errors if e["severity"] == "low")

    # 10-point scale: start at 10, deduct per severity
    score = 10.0 - (critical * 2.0) - (high * 1.0) - (medium * 0.5) - (low * 0.25)
    score = max(0.0, min(10.0, score))

    return {
        "scenario": scenario_num,
        "errors": errors,
        "counts": {"critical": critical, "high": high, "medium": medium, "low": low},
        "score": round(score, 1),
        "placeholder_count": len(placeholders),
        "paragraph_count": len(para_nums),
        "word_count": len(text.split()),
    }


def compare_drafts(pipeline_text: str, chatgpt_text: str, scenario_num: int) -> dict:
    """Compare pipeline draft vs ChatGPT-5.4 draft."""
    p_analysis = analyze_errors(pipeline_text, scenario_num)
    c_analysis = analyze_errors(chatgpt_text, scenario_num)

    # Section presence comparison
    sections_to_check = [
        "verification", "prayer", "jurisdiction", "cause of action",
        "valuation", "court fee", "legal basis", "interest",
        "list of documents", "schedule of property",
    ]
    section_comparison = {}
    for sec in sections_to_check:
        p_has = bool(re.search(sec.replace(" ", r"\s+"), pipeline_text, re.IGNORECASE))
        c_has = bool(re.search(sec.replace(" ", r"\s+"), chatgpt_text, re.IGNORECASE))
        section_comparison[sec] = {"pipeline": p_has, "chatgpt": c_has}

    return {
        "scenario": scenario_num,
        "pipeline": {
            "score": p_analysis["score"],
            "error_counts": p_analysis["counts"],
            "word_count": p_analysis["word_count"],
            "placeholder_count": p_analysis["placeholder_count"],
            "errors": p_analysis["errors"],
        },
        "chatgpt": {
            "score": c_analysis["score"],
            "error_counts": c_analysis["counts"],
            "word_count": c_analysis["word_count"],
            "placeholder_count": c_analysis["placeholder_count"],
            "errors": c_analysis["errors"],
        },
        "section_comparison": section_comparison,
        "winner": "pipeline" if p_analysis["score"] >= c_analysis["score"] else "chatgpt",
        "score_diff": round(p_analysis["score"] - c_analysis["score"], 1),
    }


# ── Draft Mode ───────────────────────────────────────────────────────────────

async def run_drafts():
    """Run pipeline on all 10 scenarios, fill Draft Agent column."""
    import openpyxl
    from app.agents.drafting_agents.drafting_graph import get_drafting_graph

    wb = openpyxl.load_workbook(str(EXCEL_PATH))
    ws = wb.active
    graph = get_drafting_graph()

    scenarios = []
    for row in ws.iter_rows(min_row=2, max_row=11, min_col=1, max_col=2, values_only=False):
        sno = row[0].value
        scenario = row[1].value
        if scenario:
            scenarios.append((sno, row[0].row, scenario.strip()))

    print("=" * 70)
    print("DRAFT BENCHMARK — 10 Civil Scenarios")
    print(f"Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"Scenarios: {len(scenarios)}")
    print("=" * 70)

    results = []
    total_start = time.perf_counter()

    for sno, excel_row, scenario in scenarios:
        print(f"\n{'─' * 70}")
        print(f"  Scenario {sno}: {scenario[:80]}...")
        print(f"{'─' * 70}")

        t0 = time.perf_counter()
        try:
            result = _as_dict(await graph.ainvoke({"user_request": scenario}))
            draft_text = _extract_draft_text(result)
            elapsed = time.perf_counter() - t0

            if not draft_text:
                draft_text = "[PIPELINE ERROR: No draft text generated]"
                print(f"  [ERROR] No draft generated ({elapsed:.1f}s)")
            else:
                # Error analysis
                analysis = analyze_errors(draft_text, sno)
                print(f"  [OK] {analysis['word_count']} words | "
                      f"Score: {analysis['score']}/10 | "
                      f"Placeholders: {analysis['placeholder_count']} | "
                      f"{elapsed:.1f}s")
                if analysis["errors"]:
                    for err in analysis["errors"][:5]:
                        print(f"    [{err['severity'].upper()}] {err['detail']}")

            # Write to Excel column C
            ws.cell(row=excel_row, column=3, value=draft_text)
            results.append({
                "scenario": sno,
                "elapsed": round(elapsed, 1),
                "word_count": len(draft_text.split()),
                "analysis": analyze_errors(draft_text, sno) if draft_text else None,
            })

        except Exception as exc:
            elapsed = time.perf_counter() - t0
            error_msg = f"[PIPELINE ERROR: {type(exc).__name__}: {exc}]"
            ws.cell(row=excel_row, column=3, value=error_msg)
            print(f"  [FAIL] {exc} ({elapsed:.1f}s)")
            results.append({"scenario": sno, "elapsed": round(elapsed, 1), "error": str(exc)})

    total_elapsed = time.perf_counter() - total_start

    # Save Excel
    wb.save(str(EXCEL_PATH))
    print(f"\n{'=' * 70}")
    print(f"  Excel saved: {EXCEL_PATH}")
    print(f"  Total time: {total_elapsed:.1f}s ({total_elapsed / 60:.1f} min)")

    # Summary
    scores = [r["analysis"]["score"] for r in results if r.get("analysis")]
    if scores:
        avg = sum(scores) / len(scores)
        print(f"  Average score: {avg:.1f}/10")
        print(f"  Min: {min(scores)}/10 | Max: {max(scores)}/10")
    print("=" * 70)

    # Save JSON report
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    report_path = OUTPUT_DIR / f"benchmark_{ts}.json"
    with open(report_path, "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False, indent=2, default=str)
    print(f"  Report: {report_path}")


# ── Compare Mode ─────────────────────────────────────────────────────────────

def run_compare():
    """Compare pipeline vs ChatGPT-5.4 after user fills column D."""
    import openpyxl

    wb = openpyxl.load_workbook(str(EXCEL_PATH))
    ws = wb.active

    print("=" * 70)
    print("DRAFT COMPARISON — Pipeline vs ChatGPT-5.4")
    print(f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 70)

    comparisons = []
    for row in ws.iter_rows(min_row=2, max_row=11, min_col=1, max_col=4, values_only=False):
        sno = row[0].value
        excel_row = row[0].row
        pipeline_text = (row[2].value or "").strip()   # Column C
        chatgpt_text = (row[3].value or "").strip()     # Column D

        if not pipeline_text or not chatgpt_text:
            print(f"\n  Scenario {sno}: SKIPPED (missing {'pipeline' if not pipeline_text else 'chatgpt'} draft)")
            continue

        comp = compare_drafts(pipeline_text, chatgpt_text, sno)
        comparisons.append(comp)

        print(f"\n{'─' * 70}")
        print(f"  Scenario {sno}: {comp['winner'].upper()} wins (diff: {comp['score_diff']:+.1f})")
        print(f"    Pipeline: {comp['pipeline']['score']}/10 | "
              f"{comp['pipeline']['word_count']} words | "
              f"{comp['pipeline']['error_counts']}")
        print(f"    ChatGPT:  {comp['chatgpt']['score']}/10 | "
              f"{comp['chatgpt']['word_count']} words | "
              f"{comp['chatgpt']['error_counts']}")

        # Section comparison
        missing_in_pipeline = [s for s, v in comp["section_comparison"].items()
                               if v["chatgpt"] and not v["pipeline"]]
        missing_in_chatgpt = [s for s, v in comp["section_comparison"].items()
                              if v["pipeline"] and not v["chatgpt"]]
        if missing_in_pipeline:
            print(f"    Pipeline missing: {', '.join(missing_in_pipeline)}")
        if missing_in_chatgpt:
            print(f"    ChatGPT missing:  {', '.join(missing_in_chatgpt)}")

        # Write Compare column (E) and Improvements column (F)
        compare_summary = (
            f"Winner: {comp['winner']} ({comp['score_diff']:+.1f})\n"
            f"Pipeline: {comp['pipeline']['score']}/10 ({comp['pipeline']['word_count']}w, "
            f"{comp['pipeline']['placeholder_count']} placeholders)\n"
            f"ChatGPT: {comp['chatgpt']['score']}/10 ({comp['chatgpt']['word_count']}w, "
            f"{comp['chatgpt']['placeholder_count']} placeholders)"
        )
        ws.cell(row=excel_row, column=5, value=compare_summary)

        # Improvements: what pipeline needs to fix
        pipeline_errors = comp["pipeline"]["errors"]
        if pipeline_errors:
            improvements = []
            for err in pipeline_errors:
                improvements.append(f"[{err['severity'].upper()}] {err['detail']}")
            if missing_in_pipeline:
                improvements.append(f"[MISSING] Sections: {', '.join(missing_in_pipeline)}")
            ws.cell(row=excel_row, column=6, value="\n".join(improvements))
        else:
            ws.cell(row=excel_row, column=6, value="No errors detected")

    # Save
    wb.save(str(EXCEL_PATH))

    # Overall summary
    if comparisons:
        p_avg = sum(c["pipeline"]["score"] for c in comparisons) / len(comparisons)
        c_avg = sum(c["chatgpt"]["score"] for c in comparisons) / len(comparisons)
        p_wins = sum(1 for c in comparisons if c["winner"] == "pipeline")
        c_wins = sum(1 for c in comparisons if c["winner"] == "chatgpt")

        print(f"\n{'=' * 70}")
        print(f"  OVERALL: Pipeline avg {p_avg:.1f}/10 vs ChatGPT avg {c_avg:.1f}/10")
        print(f"  Wins: Pipeline {p_wins} | ChatGPT {c_wins} | Tie {len(comparisons) - p_wins - c_wins}")

        # Error category breakdown
        all_p_errors = [e for c in comparisons for e in c["pipeline"]["errors"]]
        cat_counts = {}
        for e in all_p_errors:
            cat = e["category"]
            cat_counts[cat] = cat_counts.get(cat, 0) + 1
        if cat_counts:
            print(f"\n  Pipeline error breakdown:")
            for cat, count in sorted(cat_counts.items(), key=lambda x: -x[1]):
                print(f"    {cat}: {count}")
        print("=" * 70)

    # Save JSON
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    report_path = OUTPUT_DIR / f"comparison_{ts}.json"
    with open(report_path, "w", encoding="utf-8") as f:
        json.dump(comparisons, f, ensure_ascii=False, indent=2, default=str)
    print(f"  Report: {report_path}")


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Draft Benchmark Runner")
    parser.add_argument("--draft", action="store_true", default=True,
                        help="Run pipeline on all scenarios (default)")
    parser.add_argument("--compare", action="store_true",
                        help="Compare pipeline vs ChatGPT-5.4 + error analysis")
    args = parser.parse_args()

    if args.compare:
        run_compare()
    else:
        asyncio.run(run_drafts())


if __name__ == "__main__":
    main()
