"""Run pipeline on a single scenario and save to Excel.

Usage: agent_steer/Scripts/python.exe research/_run_single_scenario.py <row_number>
  row_number: Excel row (2-11 for scenarios 1-10)
"""
from __future__ import annotations

import asyncio
import sys
import time
from pathlib import Path

if sys.stdout.encoding != "utf-8":
    sys.stdout.reconfigure(encoding="utf-8")

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

EXCEL = Path(__file__).resolve().parents[1] / "docs" / "Draft_test.xlsx"


def _to_dict(v):
    if v is None:
        return {}
    if isinstance(v, dict):
        return v
    fn = getattr(v, "model_dump", None)
    return fn() if callable(fn) else {}


def _extract_draft(result: dict) -> str:
    r = _to_dict(result)
    fb = _to_dict(r.get("final_draft"))
    arts = fb.get("draft_artifacts") or []
    if arts:
        a = arts[0] if isinstance(arts[0], dict) else _to_dict(arts[0])
        t = (a.get("text") or "").strip()
        if t:
            return t
    db = _to_dict(r.get("draft"))
    da = db.get("draft_artifacts") or []
    if da:
        a = da[0] if isinstance(da[0], dict) else _to_dict(da[0])
        return (a.get("text") or "").strip()
    return ""


async def main(row: int):
    import openpyxl

    wb = openpyxl.load_workbook(str(EXCEL))
    ws = wb.active
    scenario = (ws.cell(row=row, column=2).value or "").strip()
    sno = ws.cell(row=row, column=1).value

    if not scenario:
        print(f"Row {row}: No scenario text found")
        return

    print(f"Scenario {sno} (row {row}): {scenario[:80]}...")
    print("Running pipeline...")

    from app.agents.drafting_agents.drafting_graph import get_drafting_graph

    graph = get_drafting_graph()
    t0 = time.perf_counter()
    result = await graph.ainvoke({"user_request": scenario})
    elapsed = time.perf_counter() - t0

    text = _extract_draft(result)
    wc = len(text.split()) if text else 0
    print(f"Done in {elapsed:.1f}s | Words: {wc}")

    if text:
        ws.cell(row=row, column=3, value=text)
        wb.save(str(EXCEL))
        print(f"Saved to Excel row {row}, column C")
        print(f"First 300 chars:\n{text[:300]}")
    else:
        ws.cell(row=row, column=3, value="[PIPELINE ERROR: No draft text generated]")
        wb.save(str(EXCEL))
        print("ERROR: No draft text generated")


if __name__ == "__main__":
    row = int(sys.argv[1]) if len(sys.argv) > 1 else 2
    asyncio.run(main(row))
