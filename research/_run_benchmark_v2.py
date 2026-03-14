"""Run pipeline on all 10 scenarios and save to Excel column G (Draft Agent Response 2).

Usage: agent_steer/Scripts/python.exe research/_run_benchmark_v2.py [start_row]
  start_row: Excel row to start from (default 2, range 2-11)
"""
from __future__ import annotations

import asyncio
import sys
import time
from pathlib import Path

if sys.stdout.encoding != "utf-8":
    sys.stdout.reconfigure(encoding="utf-8")

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

EXCEL = Path(__file__).resolve().parents[1] / "docs" / "Draft_test.xlsx"
TARGET_COL = 7  # Column G
TIMEOUT = 300   # 5 min max per scenario


def _to_dict(v):
    if v is None:
        return {}
    if isinstance(v, dict):
        return v
    fn = getattr(v, "model_dump", None)
    return fn() if callable(fn) else {}


def _extract_draft(result: dict) -> str:
    r = _to_dict(result)
    fb = _to_dict(r.get("final_draft"))
    arts = fb.get("draft_artifacts") or []
    if arts:
        a = arts[0] if isinstance(arts[0], dict) else _to_dict(arts[0])
        t = (a.get("text") or "").strip()
        if t:
            return t
    db = _to_dict(r.get("draft"))
    da = db.get("draft_artifacts") or []
    if da:
        a = da[0] if isinstance(da[0], dict) else _to_dict(da[0])
        return (a.get("text") or "").strip()
    return ""


async def run_one(graph, scenario: str) -> dict:
    """Run pipeline with timeout."""
    return await asyncio.wait_for(
        graph.ainvoke({"user_request": scenario}),
        timeout=TIMEOUT,
    )


async def main():
    import openpyxl

    start_row = int(sys.argv[1]) if len(sys.argv) > 1 else 2

    wb = openpyxl.load_workbook(str(EXCEL))
    ws = wb.active

    # Add header
    ws.cell(row=1, column=TARGET_COL, value="Draft Agent Response 2")
    wb.save(str(EXCEL))
    print(f"Column G = Draft Agent Response 2 | Starting from row {start_row}\n", flush=True)

    from app.agents.drafting_agents.drafting_graph import get_drafting_graph

    graph = get_drafting_graph()

    for row in range(start_row, 12):
        # Skip rows that already have valid content
        existing = (ws.cell(row=row, column=TARGET_COL).value or "").strip()
        if existing and not existing.startswith("[") and len(existing) > 200:
            print(f"Row {row}: Already has {len(existing)} chars, skipping", flush=True)
            continue

        scenario = (ws.cell(row=row, column=2).value or "").strip()
        sno = ws.cell(row=row, column=1).value

        if not scenario:
            print(f"Row {row}: No scenario text found, skipping", flush=True)
            continue

        print(f"\n{'='*60}", flush=True)
        print(f"Scenario {sno} (row {row}): {scenario[:80]}...", flush=True)
        print(f"{'='*60}", flush=True)
        print("Running pipeline...", flush=True)

        t0 = time.perf_counter()
        try:
            result = await run_one(graph, scenario)
            elapsed = time.perf_counter() - t0

            text = _extract_draft(result)
            wc = len(text.split()) if text else 0
            print(f"Done in {elapsed:.1f}s | Words: {wc}", flush=True)

            if text and "UNSUPPORTED DOMAIN" not in text:
                ws.cell(row=row, column=TARGET_COL, value=text)
                wb.save(str(EXCEL))
                print(f"Saved to Excel row {row}, column G", flush=True)
                print(f"First 200 chars:\n{text[:200]}", flush=True)
            else:
                ws.cell(row=row, column=TARGET_COL, value=f"[PIPELINE ERROR: {text[:100]}]")
                wb.save(str(EXCEL))
                print(f"ERROR: {text[:100]}", flush=True)
        except asyncio.TimeoutError:
            elapsed = time.perf_counter() - t0
            print(f"TIMEOUT after {elapsed:.1f}s (limit={TIMEOUT}s)", flush=True)
            ws.cell(row=row, column=TARGET_COL, value=f"[TIMEOUT after {elapsed:.0f}s]")
            wb.save(str(EXCEL))
        except Exception as exc:
            elapsed = time.perf_counter() - t0
            print(f"EXCEPTION after {elapsed:.1f}s: {exc}", flush=True)
            ws.cell(row=row, column=TARGET_COL, value=f"[ERROR: {exc}]")
            wb.save(str(EXCEL))

    print("\n\nAll scenarios complete!", flush=True)


if __name__ == "__main__":
    asyncio.run(main())
